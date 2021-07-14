import openpyxl

wb=openpyxl.Workbook()

wl=openpyxl.load_workbook('newaz.xlsx')


wl.create_chartsheet('SHeet2',0)

wl.remove_sheet(SHeet)
print(wl.sheetnames)